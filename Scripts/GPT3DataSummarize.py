import openai
import openpyxl
from openpyxl import load_workbook

# Load your API key from an environment variable or secret management service
openai.api_key = "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"
prompt = "Could you summarize this data? it is a collection of notes about the preformance of a frc team over their matches. Here are the notes:  "

wb = load_workbook(filename = '/Users/ben/Downloads/ScoutingExcelFormatTemplate (1).xlsx') #TODO CHANGE THIS FILE PATH TO WHERE YOU WANT THE EXCEL SHEET TO BE LOADED

for i in range(1, 10):
    ws = wb.worksheets[i]
    cell = ws.cell(12, 11)
    prompt += " " + str(cell.value) + "\n "

print("PROMPT: " + prompt)  
response = openai.Completion.create(model="text-davinci-003", prompt=prompt, temperature=0.5, max_tokens=200)

print(response.choices[0].text)