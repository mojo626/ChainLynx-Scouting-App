import pandas as pd
import numpy as np
import openpyxl
from openpyxl import load_workbook

#Device serial numbers so that we can loop through all of the devices data
deviceSerials = {"3100584a286fb200", "3100586a446ab200", "310058722270b200", "310058ba1c73b200", "310058aa2273b200", "310066705149b200"} #TODO ADD OTHER DEVICE SERIALS HERE

#Load a workbook, or create one if there isn't one here
wb = load_workbook(filename = '/Users/ben/Desktop/Coding/Documents/ChainLynxScoutingData/output.xlsx') #TODO CHANGE THIS FILE PATH TO WHERE YOU WANT THE EXCEL SHEET TO BE
ws = wb.worksheets[1]

for serial in deviceSerials:
    print(serial)
    file = open('/Users/ben/Documents/ChainLynxScoutingData/teamData' + serial + '.txt');
    if file.read().__sizeof__ == 0:
        print("File size is 0, skipping...")
        continue
    #Get JSON data
    df_json = pd.read_json('/Users/ben/Documents/ChainLynxScoutingData/teamData' + serial + '.txt', orient='records', lines=True, encoding='utf-8-sig') #TODO CHANGE THIS FILE PATH TO WHERE THE SCOUTING DATA IS BEING STORED

    for (index, row) in df_json.iterrows():
        #Check if the data is already in the excel sheet
        inSheet = False
        for row in ws.iter_rows():
            if row[12].value == df_json.get('teamNumber').values[index]:
                if  row[11].value == df_json.get('matchNumber').values[index]:
                    print("Duplicate data, skipping (match number: " + str(df_json.get('matchNumber').values[index]) + ", team number: " + str(df_json.get('teamNumber').values[index]) + ")")
                    inSheet = True
        
        #If the data isn't already in the sheet, add it
        if inSheet == False:
            ws.append(df_json.values.tolist()[index])

#Save the workbook
wb.save('/Users/ben/Desktop/Coding/Documents/ChainLynxScoutingData/output.xlsx') #CHANGE THIS FILE PATH TO WHERE YOU WANT THE EXCEL SHEET TO BE